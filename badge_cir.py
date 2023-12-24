# -*- coding: utf-8 -*-
"""

Important notes:
    
This code will generate a html file for printing circular badges. 

Before using this code, '分组' sheet in 'camp.xlsx' is needed to be revisit manualy, 
the group alloctions need to be revisited and change if needed.

The background images for badges and name are needed to be saved in 'Camp file' folder!!!

Formats and parameters of the badges can be modified in parameters region, watch out the unit!!!
     

"""


import pandas as pd
import os
from openpyxl import load_workbook
import string
upper = list(string.ascii_uppercase)  
############################# Useful funciton #################################

def name_extract(cell):
    L = cell.split(' ')
    return L[1]

###############################################################################

#####################  Parameters can be modified  ############################
 
# Badge background and dimensions in cm !!!!
badge_bg = 'bg_img.png'
badge_size = 6.7
badge_width = 6.7
badge_height = 6.7 

# Name box background and dimensions in cm !!!!!!
name_width = 6.9

# Name box relative postion to the badge, in cm !!!!!
name_margin_top = 1.45             # Postion of the name box
groupname_margin_top = 0          # Relative postion of group name
# Font of the name
name_font = "regular script"
group_font = name_font

###############################################################################

############################### Main body  ####################################

# Extract desktop path for storing 
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
newpath = desktop + '\Camp file' 

# Read the updated data
file = r'%s\camp.xlsx' % (newpath) 

# Sheet format, 0 = headers on, None = No header 
header_ = 0 

# Read the Group sheet from the 'camp' excel file 
df = pd.read_excel(file)         
df0 = load_workbook(file)
Group = df0.worksheets[0]

f = r'%s\2023.xlsx' % (newpath)
Df = pd.read_excel(f) 
Df0 = load_workbook(f)
G = Df0.worksheets[1]
tname = []

for i in range(len(Df['中文名'])):
    n = G.cell(row=i+2,column=2).value
    if n:
        tname.append(n)
    else:
        pass

# Html algorithm and design
head = '''
<head>
    <meta charset="UTF-8">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>test_cir</title>
    <style>
        * {padding:0; margin: 0;}

        @page { margin: 2cm }
        @media print  
        {
        div {
            page-break-inside: avoid;
            }
        }

        .header {
            position: fixed;
            top: 0;
            height: 20px;
            background-color: transparent;
        }
        
        .page {
            background-color: transparent;
            margin-top: 0.05cm;
            margin-left: 1cm;
            width: 21cm;
            height: 29.5cm;
            border-style: solid;
            border-width: 1px;
            border-color:transparent;
        }
        .badge{
            background-image: url(%s); 
            background-repeat: no-repeat;
            background-size: %fcm %fcm;  /* parameter */
            /*background-color: yellow;*/
            width: %fcm;
            height: %fcm;
            border-style: solid;
            border-color: transparent;
            float: left;
        }
        .name{
            width: %fcm;
            margin-top: %fcm; /* parameter */
            /*margin-left: auto;*/
            /*margin-right: auto;*/
        }
        .centre{
            padding-left: 1.30cm;
            padding-top: 0.48cm;  /* parameter */
        }
        .centre p {
            font-family: %s;
            font-size: 10.5pt;
            text-align: left;
            font-weight:bold;
            letter-spacing:4px;
            color: black;
        }
        .group_name {
            margin-top: %fcm;  /* parameter */
        }
        .group_name p {
            font-family: %s;
            font-size: 20pt;
            text-align: center;
            font-weight:bold;
            letter-spacing:4px;
            color: black;
        }
    </style>
</head>
'''% (badge_bg,badge_size,badge_size,badge_width,badge_height,name_width,\
    name_margin_top,name_font,groupname_margin_top,group_font)
    
c = '''
<div class = badge>
            <div class = name>
                <div class = centre>
                    <p>%s</p>
                </div>
                <div class = group_name >
                    <p>%s</p>
                </div>
            </div>
        </div>
''' 

# Generate badges with different names
Cc = str(c)
totalrows = len(df['组名'])

row_start = 2
Names = ''     
num = 1 
groupn = []
for i in range(totalrows):
    col = 2
    while Group.cell(row=i+2,column=col).value is not None or Group.cell(row=i+2,column=col+1).value is not None:
        if Group.cell(row=i+2,column=col).value is None:
            print(f'{upper[col-1]} passed.....')
            col += 1
        else:
            div=str(Cc)
            name = Group.cell(row=i+2,column=col).value
            name = name.replace('\n','')
            try:
                string = div%(df['组名'][i]+'组',name)
            except:
                string = div%('&nbsp',name)
            print(i+2,upper[col-1],df['组名'][i],name,num)
            col += 1
            Names += string
            num += 1 
            groupn.append(name)
    print('\n')
    
staff = []
for tn in tname:
    if tn in groupn:
        pass
    else:
        staff.append(tn)
        
#

body = '''
<body>
    <div class = page>
        %s
    </div>
</body>
</html>
''' % (Names)


# Create html file
F = open(r'%s\badge_cir.html' %(newpath),"w",encoding="utf-8")
F.write(head + body)
              
# Saving the data into the HTML file
F.close()

    
